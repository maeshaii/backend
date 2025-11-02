"""
Generate OJT Excel Template with Dropdown in Status Column
This script creates an Excel file with data validation (dropdown) for the Status field.
"""

import os
import sys
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font

def create_ojt_template_with_dropdown(output_path='ojt_template_with_status_dropdown.xlsx'):
    """Create an Excel template with dropdown for Status column"""
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "OJT Template"
    
    # Define headers based on the image shown
    headers = [
        'CTU_ID',
        'First Name',
        'Last Name',
        'Section',
        'Company',
        'Company',
        'Company',
        'Company',
        'Contact Pe',  # Contact Person
        'Position',
        'Status'
    ]
    
    # Write headers to row 1
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Make header row bold
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Create a helper sheet with status options (more reliable method)
    helper_sheet = wb.create_sheet("_StatusOptions")
    helper_sheet['A1'] = 'Ongoing'
    helper_sheet['A2'] = 'Completed'
    helper_sheet.sheet_state = 'hidden'  # Hide the helper sheet
    
    # Add data validation (dropdown) for Status column (column K, index 11)
    # Using a helper sheet reference is more reliable than comma-separated string
    status_dropdown = DataValidation(
        type="list",
        formula1="'_StatusOptions'!$A$1:$A$2",  # Reference to helper sheet
        allow_blank=False,
        showDropDown=True,
        showInputMessage=True,
        showErrorMessage=True,
        errorTitle="Invalid Status",
        error="Please select either 'Ongoing' or 'Completed'",
        promptTitle="Select Status",
        prompt="Please select a status from the dropdown list"
    )
    
    # Apply validation to Status column (K) starting from row 2 down to row 1000
    # Adjust the range as needed
    status_column_range = "K2:K1000"
    ws.add_data_validation(status_dropdown)
    status_dropdown.add(status_column_range)
    
    # Optionally add some sample data to show how it works
    sample_data = [
        ['1334287', 'Alvin', 'Dela Cruz', '4-C', '', '', '', '', '', '', 'Ongoing'],
        ['1334291', 'Jessa', 'Torres', '4-C', '', '', '', '', '', '', 'Ongoing'],
        ['1334302', 'Mark', 'Villanueva', '4-C', '', '', '', '', '', '', 'Ongoing'],
        ['1334310', 'Rhea', 'Garcia', '4-C', '', '', '', '', '', '', 'Ongoing'],
    ]
    
    # Add sample data starting from row 2
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Adjust column widths for better readability
    column_widths = {
        'A': 12,  # CTU_ID
        'B': 15,  # First Name
        'C': 15,  # Last Name
        'D': 10,  # Section
        'E': 20,  # Company columns
        'F': 20,
        'G': 20,
        'H': 20,
        'I': 15,  # Contact Pe
        'J': 15,  # Position
        'K': 15,  # Status
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Save the workbook
    wb.save(output_path)
    print(f"✅ Excel template created successfully: {output_path}")
    print(f"📋 Status column (K) has dropdown with options: Ongoing, Completed")
    print(f"📝 Sample data included. You can delete rows 2-5 if you want a blank template.")

if __name__ == '__main__':
    # Get output path from command line or use default
    output_file = sys.argv[1] if len(sys.argv) > 1 else 'ojt_template_with_status_dropdown.xlsx'
    create_ojt_template_with_dropdown(output_file)

