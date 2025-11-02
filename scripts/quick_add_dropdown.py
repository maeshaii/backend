"""
Quick script to add dropdown to Status column in an Excel file
Usage: python quick_add_dropdown.py <file_path>
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

if len(sys.argv) < 2:
    print("❌ Please provide the Excel file path")
    print("Usage: python quick_add_dropdown.py <path_to_file.xlsx>")
    print("\nExample:")
    print('  python quick_add_dropdown.py "ojt_students_update_2025-11-02_all_sections (2).xlsx"')
    sys.exit(1)

file_path = sys.argv[1]

if not os.path.exists(file_path):
    print(f"❌ File not found: {file_path}")
    sys.exit(1)

try:
    print(f"📂 Loading: {file_path}")
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Find Status column
    status_col = None
    status_col_letter = 'K'  # Default
    
    for col_idx in range(1, min(20, ws.max_column + 1)):  # Check first 20 columns
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value and 'status' in str(cell_value).lower():
            status_col_letter = ws.cell(row=1, column=col_idx).column_letter
            status_col = col_idx
            break
    
    print(f"✅ Status column found at: {status_col_letter}")
    
    # Remove existing validations for this column
    if ws.data_validations:
        ws.data_validations.dataValidation = [
            dv for dv in ws.data_validations.dataValidation
            if status_col_letter not in str(dv.ranges)
        ]
    
    # Create helper sheet
    helper_name = "_StatusOptions"
    if helper_name not in [s.title for s in wb.worksheets]:
        helper = wb.create_sheet(helper_name)
        helper['A1'] = 'Ongoing'
        helper['A2'] = 'Completed'
        helper.sheet_state = 'hidden'
    
    # Create dropdown
    dropdown = DataValidation(
        type="list",
        formula1=f"'{helper_name}'!$A$1:$A$2",
        allow_blank=False,
        showDropDown=True
    )
    
    max_row = max(ws.max_row, 100)
    dropdown.add(f"{status_col_letter}2:{status_col_letter}{max_row}")
    ws.add_data_validation(dropdown)
    
    # Save (overwrite original)
    wb.save(file_path)
    print(f"✅ Dropdown added to Status column!")
    print(f"📋 Options: Ongoing, Completed")
    print(f"📝 Applied to {status_col_letter}2:{status_col_letter}{max_row}")
    print(f"\n⚠️ IMPORTANT: Open this file in Microsoft Excel DESKTOP APP (not web browser) to see the dropdown!")
    
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

