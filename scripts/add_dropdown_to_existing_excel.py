"""
Add Dropdown to Status Column in Existing Excel File
This script adds data validation dropdown to the Status column of an existing Excel file.
"""

import os
import sys
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

def add_dropdown_to_existing_file(input_path, output_path=None):
    """Add dropdown validation to Status column in existing Excel file"""
    
    if not os.path.exists(input_path):
        print(f"❌ Error: File not found: {input_path}")
        return False
    
    try:
        # Load the existing workbook
        print(f"📂 Loading file: {input_path}")
        wb = load_workbook(input_path)
        ws = wb.active
        
        # Find the Status column (column K or find it by header)
        status_col = None
        status_col_letter = None
        
        # Check row 1 for "Status" header
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            if cell_value and str(cell_value).strip().lower() == 'status':
                status_col = col_idx
                status_col_letter = ws.cell(row=1, column=col_idx).column_letter
                print(f"✅ Found Status column at: {status_col_letter} ({status_col})")
                break
        
        if not status_col:
            # If not found, try column K (index 11)
            status_col = 11
            status_col_letter = 'K'
            print(f"⚠️ Status header not found, using column K by default")
        
        # Remove any existing data validation from Status column
        if ws.data_validations:
            to_remove = []
            for dv in ws.data_validations.dataValidation:
                # Check if this validation applies to Status column
                if status_col_letter in str(dv.ranges):
                    to_remove.append(dv)
            for dv in to_remove:
                ws.data_validations.dataValidation.remove(dv)
        
        # Create a helper sheet with status options (more reliable method)
        helper_sheet_name = "_StatusOptions"
        if helper_sheet_name not in [s.title for s in wb.worksheets]:
            helper_sheet = wb.create_sheet(helper_sheet_name)
            helper_sheet['A1'] = 'Ongoing'
            helper_sheet['A2'] = 'Completed'
            helper_sheet.sheet_state = 'hidden'  # Hide the helper sheet
        else:
            helper_sheet = wb[helper_sheet_name]
        
        # Create new data validation dropdown using helper sheet reference
        status_dropdown = DataValidation(
            type="list",
            formula1=f"'{helper_sheet_name}'!$A$1:$A$2",  # Reference to helper sheet
            allow_blank=False,
            showDropDown=True,
            showInputMessage=True,
            showErrorMessage=True,
            errorTitle="Invalid Status",
            error="Please select either 'Ongoing' or 'Completed'",
            promptTitle="Select Status",
            prompt="Please select a status from the dropdown list: Ongoing or Completed"
        )
        
        # Apply validation to Status column starting from row 2
        max_row = ws.max_row if ws.max_row > 1000 else 1000
        status_column_range = f"{status_col_letter}2:{status_col_letter}{max_row}"
        
        ws.add_data_validation(status_dropdown)
        status_dropdown.add(status_column_range)
        
        # Determine output path
        if output_path is None:
            base_name = os.path.splitext(input_path)[0]
            output_path = f"{base_name}_with_dropdown.xlsx"
        
        # Save the workbook
        wb.save(output_path)
        print(f"✅ Excel file saved with dropdown: {output_path}")
        print(f"📋 Status column ({status_col_letter}) now has dropdown with options: Ongoing, Completed")
        print(f"📝 Applied to rows 2-{max_row}")
        return True
        
    except Exception as e:
        print(f"❌ Error processing file: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python add_dropdown_to_existing_excel.py <input_file.xlsx> [output_file.xlsx]")
        print("\nExample:")
        print("  python add_dropdown_to_existing_excel.py ojt_students_update.xlsx")
        print("  python add_dropdown_to_existing_excel.py input.xlsx output.xlsx")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    add_dropdown_to_existing_file(input_file, output_file)

